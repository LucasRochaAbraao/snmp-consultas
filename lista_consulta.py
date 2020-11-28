#!/usr/bin/env python3
# coding=utf-8
r"""
    Esse script consulta o sinal das ONUs passadas no arquivo
    'descs.xlsx' e retorna no formato: DESC | STATUS | SINAL

                             } (
Author:  Lucas Rocha Abraão (   ) )
Date:    20/06/2020          ) { (
Versão:  2.0              ___|___)__
License: GNU GPLv3     .-'---------|
                      ( C|/\/\/\/\/|
                       '-./\/\/\/\/|
                         '_________'
                          '-------'
"""

import configparser
import openpyxl
import pathlib
import time
import sys
from snmp import py_snmp

class cor:
    VM  = '\033[91m' # vermelho
    VD  = '\033[92m' # verde
    LR  = '\033[93m' # laranja
    AZ  = '\033[94m' # azul
    UN  = '\033[4m'  # underline
    NG  = '\033[1m'  # negrito
    FIM = '\033[0m'

def load_xlsx(arquivo):
    workbook  = openpyxl.load_workbook(arquivo)
    worksheet = workbook.active
    dados = {}
    for col in range(1, worksheet.max_column+1):
        col_atual = []
        for row in range(1, worksheet.max_row + 1):
            col_atual.append(worksheet.cell(row=row, column=col).value)
        dados[col_atual[0]] = [col, col_atual[1:]] # atualiza o dict com 'olt': [coluna atual, descs]
    # como passar a posição da coluna? Para ser escrito ao lado o sinal.
    return dados

def write_xlsx(olt, dados, col):
    workbook  = openpyxl.load_workbook('descs/descs_online.xlsx')
    worksheet = workbook.active
    if len(dados) < 1:
        print("Não foi possível coletar os dados")
        sys.exit()
    worksheet.cell(row=1, column=col,   value=olt[0])
    worksheet.cell(row=1, column=col+1, value='Sinais')
    for row, item in enumerate(dados):
        worksheet.cell(row=row+2, column=col,   value=item[0]) # item = ('desc', 'online', 'sinal')
        worksheet.cell(row=row+2, column=col+1, value=item[2])
    workbook.save('descs/descs_online.xlsx')

def get_olts(file_name, item):
    config = configparser.ConfigParser()
    config.read(str(pathlib.Path(__file__).parent.absolute()) + f'/{file_name}.ini')
    return config[item]

def consulta(olt, descs):
    print("\n=================================\n")
    print(f"\n --> Verificando os clientes [{' '.join(descs)}] da olt de {cor.NG}{cor.AZ}{cor.UN}{olt[0].upper()}{cor.FIM}.\n")
    print(">  descs....", end="\t")
    sys.stdout.flush() # para sair o print anterior, pois sem o \n no end= ele aguarda o próximo print()
    start_desc = time.time()
    descs_all = py_snmp.descricao(olt[1], 'qn31415926')
    fim_desc = time.time() - start_desc
    if fim_desc > 60:
        fim_desc = f'{int(fim_desc/60)}m{int(fim_desc%60)}s'
    else:
        fim_desc = f'{time.time() - start_desc:.1f}s'
    print(f"{cor.VD}✓ {fim_desc}{cor.FIM}")

    print(">  status...", end="\t")
    sys.stdout.flush()
    start_status = time.time()
    status_all = py_snmp.status(olt[1], 'qn31415926')
    fim_status = int(time.time() - start_status)
    if fim_status > 60:
        fim_status = f'{int(fim_status/60)}m{int(fim_status%60)}s'
    else:
        fim_status = f'{time.time() - start_status:.1f}s'
    print(f"{cor.VD}✓ {fim_status}{cor.FIM}")

    print(">  sinais...", end="\t")
    sys.stdout.flush()
    start_sinais = time.time()
    sinais_all = py_snmp.potencia(olt[1], 'qn31415926')
    fim_sinais = int(time.time() - start_sinais)
    if fim_sinais > 60:
        fim_sinais = f'{int(fim_sinais/60)}m{int(fim_sinais%60)}s'
    else:
        fim_sinais = f'{time.time() - start_sinais:.1f}s'
    print(f"{cor.VD}✓ {fim_sinais}{cor.FIM}")
    
    time.sleep(0.5)
    #print(len(descs_all), len(status_all), len(sinais_all))
    resultado = []
    for desc, status, sinal in zip(descs_all, status_all, sinais_all):
        if desc in descs:
            #status = f'{cor.VD}{status}{cor.FIM}'
            if sinal == '21474836.47':
            #    status = f'{cor.VM}{status}{cor.FIM}'
                sinal = 'offline'
            resultado.append([desc, status, sinal])
    
    # Verificar se todas foram encontradas...
    return resultado
    # Apresentação final - passar pra outra função, essa deve ser apenas consulta()
    # retornar valores pra serem gravados.
#    print("\n---------------------------------")
#    print(f'  {cor.UN}{cor.NG}DESC{cor.FIM}\t|   {cor.UN}{cor.NG}STATUS{cor.FIM}\t|   {cor.UN}{cor.NG}SINAL{cor.FIM}\n')
#    for res in resultado:
#        print(f'  {res[0]}\t    {res[1]}\t   {res[2]}')
    

if __name__ == '__main__':

    arquivo = 'descs/descs.xlsx'
    DADOS = load_xlsx(arquivo)

    olts_parser = get_olts('snmp/config', 'OLTS')
    todas_olts = {}
    for olt in DADOS.keys():
        if olt == None:
            continue
        todas_olts[olt.lower()] = olts_parser[olt].split()[0]

#    resultado_consultas = {}
    inicio = time.time()
    print(DADOS)
    coluna = 1
    for dado in DADOS.items():
        if dado[0] == None: # ele pode pegar uma coluna vazia, então verifico se não tem olt e pulo
            continue
        #coluna = dado[1][0]
        descs = [str(desc) for desc in dado[1][1] if desc != None]
        olt_atual = (dado[0].lower(), todas_olts[dado[0].lower()])
        retorno_consulta = consulta(olt_atual, descs)
        if len(retorno_consulta) > 1:
            write_xlsx(olt_atual, retorno_consulta, coluna)
            coluna += 2
        else:
            print("Não houve retorno")
            continue
#        resultado_consultas[dado[0]] = consulta(olt_atual, descs)
#    print(resultado_consultas) # ('olt', '10.80.80.x') (col, '1234', '5678', etc)
    fim = time.time() - inicio

    print(f'\nTempo total de consulta: {cor.LR}{int(fim/60)}m{int(fim%60)}s{cor.FIM}\n')
